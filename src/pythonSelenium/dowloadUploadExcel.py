import time

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait
import openpyxl

def update_excel_date(file_path,fruit_name,columnName,newValue):
    Dict = {}
    book = openpyxl.load_workbook(file_path)
    sheet = book.active

    for i in range(1, sheet.max_column+1):
        if sheet.cell(row=1, column=i).value == columnName:
            Dict["col"] = i

    for i in range(1, sheet.max_row+1):
        for j in range(1, sheet.max_column+1):
            if sheet.cell(row=i, column=j).value == fruit_name:
                Dict["row"] = i

    sheet.cell(row=Dict["row"],column=Dict["col"]).value =   newValue
    book.save(file_path)

file_path = "C:\\Users\\Arunr\\Downloads\\download.xlsx"
fruit_name = "Apple"
columnName = "price"
newValue = "999"

driver = webdriver.Chrome()
driver.maximize_window()
driver.implicitly_wait(5)

driver.get("https://rahulshettyacademy.com/upload-download-test/")

#Download the excel
driver.find_element(By.ID,"downloadButton").click()

#Edit the excel
update_excel_date(file_path,fruit_name,columnName,newValue)

#upload the excel
file_input = driver.find_element(By.CSS_SELECTOR,"input[type='file']")
time.sleep(2)
file_input.send_keys(file_path)

wait = WebDriverWait(driver,5)
toast_locator = (By.CSS_SELECTOR,".Toastify__toast-body div:nth-child(2)")
wait.until(expected_conditions.visibility_of_element_located(toast_locator))
print(driver.find_element(*toast_locator).text)
priceColum = driver.find_element(By.XPATH,"//div[text()='Price']/parent::div").get_attribute("data-column-id")
actual_price = driver.find_element(By.XPATH,"//div[text()='"+fruit_name+"']/parent::div/parent::div/div[@id='cell-"+priceColum+"-undefined']").text
assert actual_price == newValue